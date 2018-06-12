import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from sklearn.cluster import KMeans
from sklearn.datasets import make_blobs
from openpyxl import load_workbook
from openpyxl import Workbook
from sklearn.neighbors import LocalOutlierFactor

plt.rcParams['figure.figsize'] = (16, 9)
dataused = 13
clusters = 7

#load  workbook
wb = load_workbook('genes.xlsx')
ws = wb['Sheet1']

#fill data matrix
d = np.zeros((3583,dataused))
for i in range(2,2+dataused):
    for j in range(2,3585):
        d[j-2][i-2] = ws.cell(row=j,column=i).value

#fill parallel names list
names = []
for i in range(2,3585):
    names.append(ws.cell(row = i, column = 1).value)
#outlier removal
clf = LocalOutlierFactor(n_neighbors=20)
y_pred = clf.fit_predict(d)
iterator = 0
while (i<len(y_pred)):
    if y_pred[i] == -1:
        d = np.delete(d, (i), axis = 0)
        names.pop(i)
        y_pred = np.delete(y_pred, (i), axis = 0)
        i-=1
    i+=1

print(len(y_pred))
print(len(names))
print(len(d))

# Initializing KMeans
kmeans = KMeans(n_clusters=clusters)
# Fitting with inputs
kmeans = kmeans.fit(d)
# Predicting the clusters
labels = kmeans.predict(d)
# Getting the cluster centers
C = kmeans.cluster_centers_

#plot cost function
e = np.zeros((20,1))
f = np.zeros((20,1))
for i in range (1, 21):
    # Initializing KMeans
    kmeans = KMeans(n_clusters=i)
    # Fitting with inputs
    kmeans = kmeans.fit(d)
    # Predicting the clusters
    labelstest = kmeans.predict(d)
    # Getting the cluster centers
    e[i-1] = kmeans.inertia_
    f[i-1] = i

fig = plt.figure()
ax = fig.add_subplot(111)
ax.plot(f,e)
#ax.set_yscale('log')
fig = plt.show()

##for item in labels:
##    print(item)

#create new sheet
sf = Workbook()
sf.active.title = "cluster 1"
for i in range(2,clusters+1):
    sf.create_sheet("cluster "+str(i))
sheetnum = 0

#fill sheet
for sheet in sf:
    varindex = 1
    for i in range (0, labels.size):
        if labels[i] == sheetnum:
            sheet.cell(row = varindex, column = 1).value = names[i]
            for j in range(2,2+dataused):
                sheet.cell(row = varindex, column = j).value = d[i][j-2]
            varindex +=1
    sheetnum+=1
sf.save('KmeansNoOutliers.xlsx')       
            

import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from sklearn.cluster import KMeans
from sklearn.datasets import make_blobs
from openpyxl import load_workbook
from sklearn.neighbors import LocalOutlierFactor
from sklearn.cluster import MeanShift, estimate_bandwidth
from sklearn.datasets.samples_generator import make_blobs
from itertools import cycle

#open workbook
wb = load_workbook('genes.xlsx')
ws = wb['Sheet1']

#initialize and fill matrix
d = np.zeros((3584,3))
for i in range(2,5):
    for j in range(2,3585):
        d[j-2][i-2] = ws.cell(row=j,column=i).value

##clf = LocalOutlierFactor(n_neighbors=20)
##y_pred = clf.fit_predict(d)
##iterator = 0
##while (i<len(y_pred)):
##    if y_pred[i] == -1:
##        d = np.delete(d, (i), axis = 0)
##        y_pred = np.delete(y_pred, (i), axis = 0)
##        i-=1
##    i+=1

#plot initial data
plt.rcParams['figure.figsize'] = (16, 9)
fig = plt.figure()
ax = Axes3D(fig)
ax.scatter(d[:, 0], d[:, 1], d[:, 2])
##ax.set_xlim3d(0,1)
##ax.set_ylim3d(0,1)
##ax.set_zlim3d(0,1)
fig = plt.show()


bandwidth = estimate_bandwidth(d, quantile=0.2, n_samples=500)
ms = MeanShift(bandwidth=bandwidth, bin_seeding=True)
ms.fit(d)
labels = ms.labels_
C = ms.cluster_centers_
labels_unique = np.unique(labels)
n_clusters_ = len(labels_unique)
print("number of estimated clusters : %d" % n_clusters_)

#plot clusters
plt.rcParams['figure.figsize'] = (16, 9)
fig = plt.figure()
ax = Axes3D(fig)
ax.scatter(d[:, 0], d[:, 1], d[:, 2])
ax.scatter(C[:, 0], C[:, 1], C[:, 2], marker='*', c='#050505', s=500)
fig = plt.show()




    


import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from sklearn.cluster import KMeans
from sklearn.datasets import make_blobs
from openpyxl import load_workbook
from openpyxl import Workbook
from sklearn.neighbors import LocalOutlierFactor
from scipy.cluster.hierarchy import dendrogram, linkage, fcluster

plt.rcParams['figure.figsize'] = (16, 9)
dataused = 13
clusters = 7

#load  workbook
wb = load_workbook('genes.xlsx')
ws = wb['Sheet1']

#fill data matrix
d = np.zeros((3583,dataused))
for i in range(2,2+dataused):
    for j in range(2,3585):
        d[j-2][i-2] = ws.cell(row=j,column=i).value

#fill parallel names list
names = []
for i in range(2,3585):
    names.append(ws.cell(row = i, column = 1).value)
    
Z = linkage(d, 'ward')


# calculate full dendrogram
plt.figure(figsize=(25, 10))
plt.title('Hierarchical Clustering Dendrogram')
plt.xlabel('sample index')
plt.ylabel('distance')
dendrogram(
    Z,
    truncate_mode='lastp',  # show only the last p merged clusters
    p=100,  # show only the last p merged clusters
    leaf_rotation=90.,
    leaf_font_size=12.,
    show_contracted=True,  # to get a distribution impression in truncated branches

)
plt.show()

arr = fcluster(Z, clusters, criterion='maxclust')

#create new sheet
sf = Workbook()
sf.active.title = "cluster 1"
for i in range(2,clusters+1):
    sf.create_sheet("cluster "+str(i))
sheetnum = 1

for sheet in sf:
    varindex = 1
    for i in range (0, arr.size):
        if arr[i] == sheetnum:
            sheet.cell(row = varindex, column = 1).value = names[i]
            for j in range(2,2+dataused):
                sheet.cell(row = varindex, column = j).value = d[i][j-2]
            varindex +=1
    sheetnum+=1
sf.save('hierarchcluster.xlsx')       

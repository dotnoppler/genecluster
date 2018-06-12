import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from sklearn.cluster import KMeans
from sklearn.datasets import make_blobs
from openpyxl import load_workbook
from sklearn.neighbors import LocalOutlierFactor

#open workbook
wb = load_workbook('genes.xlsx')
ws = wb['Sheet1']

#initialize and fill matrix
d = np.zeros((3584,3))
for i in range(2,5):
    for j in range(2,3585):
        d[j-2][i-2] = ws.cell(row=j,column=i).value

##clf = LocalOutlierFactor(n_neighbors=20)
##y_pred = clf.fit_predict(d)
##iterator = 0
##while (i<len(y_pred)):
##    if y_pred[i] == -1:
##        d = np.delete(d, (i), axis = 0)
##        y_pred = np.delete(y_pred, (i), axis = 0)
##        i-=1
##    i+=1

#plot initial data
plt.rcParams['figure.figsize'] = (16, 9)
fig = plt.figure()
ax = Axes3D(fig)
ax.scatter(d[:, 0], d[:, 1], d[:, 2])
##ax.set_xlim3d(0,1)
##ax.set_ylim3d(0,1)
##ax.set_zlim3d(0,1)
fig = plt.show()


#plotting cost function
e = np.zeros((20,1))
f = np.zeros((20,1))
for i in range (1, 21):
    # Initializing KMeans
    kmeans = KMeans(n_clusters=i)
    # Fitting with inputs
    kmeans = kmeans.fit(d)
    # Predicting the clusters
    labels = kmeans.predict(d)
    # Getting the cluster centers
    e[i-1] = kmeans.inertia_
    f[i-1] = i
    
fig = plt.figure()
ax = fig.add_subplot(111)
ax.plot(f,e)
#ax.set_yscale('log')
fig = plt.show()

# Initializing KMeans
kmeans = KMeans(n_clusters=5)
# Fitting with inputs
kmeans = kmeans.fit(d)
# Predicting the clusters
labels = kmeans.predict(d)
# Getting the cluster centers
C = kmeans.cluster_centers_

#plot clusters
plt.rcParams['figure.figsize'] = (16, 9)
fig = plt.figure()
ax = Axes3D(fig)
ax.scatter(d[:, 0], d[:, 1], d[:, 2], c=kmeans.labels_.astype(float))
ax.scatter(C[:, 0], C[:, 1], C[:, 2], marker='*', c='#050505', s=500)
fig = plt.show()

#plot clusters
plt.rcParams['figure.figsize'] = (16, 9)
fig = plt.figure()
ax = Axes3D(fig)
ax.scatter(d[:, 0], d[:, 1], d[:, 2], c=kmeans.labels_.astype(float))
ax.scatter(C[:, 0], C[:, 1], C[:, 2], marker='*', c='#050505', s=500)
ax.set_xlim3d(0,5)
ax.set_ylim3d(0,5)
ax.set_zlim3d(0,5)
fig = plt.show()


    

